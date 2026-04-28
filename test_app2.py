import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
from datetime import datetime

# -------------------------------------------------------------------
# PAGE CONFIG
# -------------------------------------------------------------------
st.set_page_config(page_title="Scalpel AI – Deterministic Accounting Engine", layout="wide")
st.title("🔪 Scalpel AI")
st.markdown("### Deterministic Finance Operating System")


# -------------------------------------------------------------------
# DATA NORMALISATION
# -------------------------------------------------------------------
def normalize_columns(df):
    df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
    rename_map = {
        'account_description': 'account_description',
        'accountname': 'account_description',
        'account_desc': 'account_description',
        'effective_date': 'effective_date',
        'effectivedate': 'effective_date',
        'entered_date': 'entered_date',
        'entereddate': 'entered_date',
        'transaction': 'transaction',
        'txnid': 'transaction',
        'glid': 'transaction',
        'debit': 'debit',
        'debits': 'debit',
        'credit': 'credit',
        'credits': 'credit',
        'memo': 'memo',
        'description': 'memo'
    }
    for old, new in rename_map.items():
        if old in df.columns and new not in df.columns:
            df.rename(columns={old: new}, inplace=True)
    return df


@st.cache_data
def load_and_validate(uploaded_file):
    df = pd.read_excel(uploaded_file)
    df = normalize_columns(df)

    # Ensure required columns
    required = ['account', 'debit', 'credit']
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}. Please ensure file has account, debit, credit.")
        return None, False

    # Fill missing optionals
    if 'account_description' not in df.columns:
        df['account_description'] = df['account'].astype(str)
    if 'effective_date' not in df.columns and 'entered_date' in df.columns:
        df['effective_date'] = df['entered_date']
    if 'effective_date' not in df.columns:
        df['effective_date'] = pd.NaT
    if 'transaction' not in df.columns:
        df['transaction'] = df.index.astype(str)
    if 'memo' not in df.columns:
        df['memo'] = ''

    # Convert numeric
    df['debit'] = pd.to_numeric(df['debit'], errors='coerce').fillna(0)
    df['credit'] = pd.to_numeric(df['credit'], errors='coerce').fillna(0)

    # Global balance check
    total_debits = df['debit'].sum()
    total_credits = df['credit'].sum()
    if abs(total_debits - total_credits) > 0.01:
        st.error(f"❌ Global imbalance: Debits = {total_debits:,.2f}, Credits = {total_credits:,.2f}")
        return None, False
    else:
        st.success(f"✅ Global balance verified (Debits = Credits = {total_debits:,.2f})")

    return df, True


# -------------------------------------------------------------------
# ACCOUNT CLASSIFICATION (CORRECTED)
# -------------------------------------------------------------------
def classify_account(acct_num, acct_desc):
    acct_str = str(acct_num)
    desc_lower = acct_desc.lower()

    # Revenue: 4xxx or sales/revenue keywords
    if acct_str.startswith('4') or any(k in desc_lower for k in
                                       ['sales', 'revenue', 'consulting', 'drafting', 'surveying', 'service',
                                        'freight revenue']):
        return "Revenue"

    # COGS: 5xxx that are material/inventory related OR explicitly COGS
    if acct_str.startswith('5'):
        if any(k in desc_lower for k in
               ['material', 'cogs', 'purchase discounts', 'adjustment write-off', 'equipment rental', 'freight expense',
                'item assembly']):
            return "COGS"
        else:
            return "OpEx"  # e.g., 54100 Wages & Salaries

    # Asset: 1xxxx or asset keywords
    if acct_str.startswith('1') or any(k in desc_lower for k in
                                       ['cash', 'bank', 'receivable', 'inventory', 'prepaid', 'accum amort', 'vehicles',
                                        'drywall', 'hardware', 'lumber', 'roofing']):
        return "Asset"

    # Liability: 2xxxx or payable keywords
    if acct_str.startswith('2') or 'payable' in desc_lower:
        return "Liability"

    # Equity: 3xxxx
    if acct_str.startswith('3'):
        return "Equity"

    return "Other"


def add_classification(df):
    unique_acc = df[['account', 'account_description']].drop_duplicates()
    unique_acc['category'] = unique_acc.apply(lambda r: classify_account(r['account'], r['account_description']),
                                              axis=1)
    return df.merge(unique_acc[['account', 'category']], on='account', how='left')


# -------------------------------------------------------------------
# FINANCIAL STATEMENTS (DETAILED)
# -------------------------------------------------------------------
def generate_trial_balance(df):
    tb = df.groupby(['account', 'account_description']).agg({'debit': 'sum', 'credit': 'sum'}).reset_index()
    tb['balance'] = tb['debit'] - tb['credit']
    tb['category'] = tb.apply(lambda r: classify_account(r['account'], r['account_description']), axis=1)
    return tb


def generate_detailed_pl(df, tb):
    """
    Returns:
      - pl_summary: DataFrame with Revenue, COGS, Gross Profit, OpEx, Net Income
      - revenue_details: DataFrame of each revenue account (credit - debit)
      - cogs_details: DataFrame of each COGS account (debit only → because COGS is normally debit)
      - opex_details: DataFrame of each OpEx account (debit - credit), with subcategory
    """
    # Revenue: credit - debit
    rev_df = tb[tb['category'] == 'Revenue'].copy()
    rev_df['amount'] = rev_df['credit'] - rev_df['debit']
    rev_details = rev_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_revenue = rev_details['amount'].sum()

    # COGS: debit only (cost of goods sold is a debit; credits are returns/reclassifications that we exclude from P&L)
    cogs_df = tb[tb['category'] == 'COGS'].copy()
    cogs_df['amount'] = cogs_df['debit']  # no subtraction of credits – those are inventory reclassifications
    cogs_details = cogs_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_cogs = cogs_details['amount'].sum()

    gross_profit = total_revenue - total_cogs

    # OpEx: debit - credit
    opex_df = tb[tb['category'] == 'OpEx'].copy()
    opex_df['amount'] = opex_df['debit'] - opex_df['credit']
    opex_details = opex_df[['account_description', 'amount']].sort_values('amount', ascending=False)

    # Add subcategory based on keywords
    def subcategory(desc):
        d = desc.lower()
        if any(k in d for k in
               ['wage', 'salary', 'ei', 'cpp', 'qpp', 'wcb', 'eht', 'qpip', 'qhsf', 'rrsp', 'union', 'medical',
                'subcontractor']):
            return "Payroll & Benefits"
        if any(k in d for k in
               ['amortization', 'amort', 'insurance', 'rent', 'repair', 'maintenance', 'bell', 'utilities']):
            return "Facilities & Operations"
        if any(k in d for k in
               ['accounting', 'legal', 'advertising', 'promotion', 'courier', 'postage', 'interest', 'bank',
                'office supplies', 'travel', 'amex', 'licenses']):
            return "Professional & Admin"
        return "Other Operating Expenses"

    opex_details['subcategory'] = opex_details['account_description'].apply(subcategory)

    total_opex = opex_details['amount'].sum()
    net_income = gross_profit - total_opex

    # Summary DataFrame
    summary = pd.DataFrame([
        ("Revenue", total_revenue),
        ("COGS", total_cogs),
        ("Gross Profit", gross_profit),
        ("Operating Expenses", total_opex),
        ("Net Income", net_income)
    ], columns=["Line Item", "Amount"])

    return summary, rev_details, cogs_details, opex_details, net_income


def generate_detailed_balance_sheet(tb, net_income):
    """
    Returns:
      - assets_details: DataFrame of asset accounts (normal debit balance)
      - liabilities_details: DataFrame of liability accounts (normal credit balance)
      - equity_details: DataFrame of equity accounts + net income
    """
    asset_df = tb[tb['category'] == 'Asset'].copy()
    asset_df['amount'] = asset_df['balance']  # balance = debit-credit (positive for assets)
    assets_details = asset_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_assets = assets_details['amount'].sum()

    liability_df = tb[tb['category'] == 'Liability'].copy()
    liability_df['amount'] = -liability_df['balance']  # because credit balance is negative in balance column
    liabilities_details = liability_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_liabilities = liabilities_details['amount'].sum()

    # Equity: from trial balance (if any) + net income
    equity_df = tb[tb['category'] == 'Equity'].copy()
    equity_df['amount'] = equity_df['balance']  # credit balance in equity is positive
    total_equity_from_tb = equity_df['amount'].sum()
    total_equity = total_equity_from_tb + net_income
    # Create a simple equity line item
    equity_items = []
    if total_equity_from_tb != 0:
        equity_items.append(("Equity from TB", total_equity_from_tb))
    equity_items.append(("Net Income", net_income))
    equity_items.append(("Total Equity", total_equity))
    equity_details = pd.DataFrame(equity_items, columns=["Account Description", "Amount"])

    return assets_details, liabilities_details, equity_details, total_assets, total_liabilities, total_equity


def generate_monthly_analysis(df):
    if 'effective_date' not in df.columns or df['effective_date'].isnull().all():
        return None
    df['date'] = pd.to_datetime(df['effective_date'], errors='coerce')
    if df['date'].isnull().all():
        return None
    df['year_month'] = df['date'].dt.to_period('M')
    df_class = add_classification(df)
    monthly = df_class.groupby('year_month').apply(
        lambda x: pd.Series({
            'Revenue': x[x['category'] == 'Revenue']['credit'].sum() - x[x['category'] == 'Revenue']['debit'].sum(),
            'COGS': x[x['category'] == 'COGS']['debit'].sum(),  # debit only
            'OpEx': x[x['category'] == 'OpEx']['debit'].sum() - x[x['category'] == 'OpEx']['credit'].sum()
        })
    ).reset_index()
    monthly['Gross Profit'] = monthly['Revenue'] - monthly['COGS']
    monthly['Net Income'] = monthly['Gross Profit'] - monthly['OpEx']
    monthly['Gross Margin %'] = (monthly['Gross Profit'] / monthly['Revenue'] * 100).round(1)
    monthly['Net Margin %'] = (monthly['Net Income'] / monthly['Revenue'] * 100).round(1)
    monthly['year_month'] = monthly['year_month'].astype(str)
    return monthly


# -------------------------------------------------------------------
# LEDGER VIEW (PER ACCOUNT)
# -------------------------------------------------------------------
def get_account_ledger(df, account_num):
    acc_df = df[df['account'] == account_num].copy()
    if acc_df.empty:
        return pd.DataFrame()
    # sort by date
    date_col = 'effective_date' if 'effective_date' in acc_df.columns and acc_df[
        'effective_date'].notna().any() else 'entered_date'
    if date_col in acc_df.columns:
        acc_df['date'] = pd.to_datetime(acc_df[date_col], errors='coerce')
        acc_df = acc_df.sort_values('date')
    else:
        acc_df['date'] = pd.NaT
    show_cols = ['date', 'memo', 'debit', 'credit']
    for c in show_cols:
        if c not in acc_df.columns:
            acc_df[c] = ''
    ledger = acc_df[show_cols].copy()
    ledger['debit'] = ledger['debit'].fillna(0)
    ledger['credit'] = ledger['credit'].fillna(0)
    total_debit = ledger['debit'].sum()
    total_credit = ledger['credit'].sum()
    total_row = pd.DataFrame([['TOTAL', '', total_debit, total_credit]], columns=show_cols)
    ledger = pd.concat([ledger, total_row], ignore_index=True)
    return ledger


# -------------------------------------------------------------------
# EXCEL DOWNLOAD (DETAILED, STYLED)
# -------------------------------------------------------------------
def download_detailed_excel(tb, pl_summary, rev_details, cogs_details, opex_details, assets_details,
                            liabilities_details, equity_details, monthly_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        tb.to_excel(writer, sheet_name='Trial Balance', index=False)
        pl_summary.to_excel(writer, sheet_name='P&L Summary', index=False)
        rev_details.to_excel(writer, sheet_name='Revenue Details', index=False)
        cogs_details.to_excel(writer, sheet_name='COGS Details', index=False)
        opex_details.to_excel(writer, sheet_name='OpEx Details', index=False)
        assets_details.to_excel(writer, sheet_name='Assets', index=False)
        liabilities_details.to_excel(writer, sheet_name='Liabilities', index=False)
        equity_details.to_excel(writer, sheet_name='Equity', index=False)
        if monthly_df is not None:
            monthly_df.to_excel(writer, sheet_name='Monthly Analysis', index=False)
        # Basic styling (bold headers, auto-width)
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 30)
    return output.getvalue()


# -------------------------------------------------------------------
# MAIN APP
# -------------------------------------------------------------------
def main():
    uploaded_file = st.file_uploader("📂 Upload General Ledger (Excel)", type=["xlsx", "xls"])
    if uploaded_file is None:
        st.info("Please upload your general ledger file to start.")
        return

    df, is_valid = load_and_validate(uploaded_file)
    if not is_valid or df is None:
        st.stop()

    # Add classification
    df_class = add_classification(df)

    # Generate reports
    tb = generate_trial_balance(df)
    pl_summary, rev_details, cogs_details, opex_details, net_income = generate_detailed_pl(df_class, tb)
    assets_details, liabilities_details, equity_details, total_assets, total_liabilities, total_equity = generate_detailed_balance_sheet(
        tb, net_income)
    monthly = generate_monthly_analysis(df)

    # -------------------------------------------------------------------
    # UI TABS
    # -------------------------------------------------------------------
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📒 Account Ledgers", "📊 Trial Balance", "📈 P&L (Detailed)",
        "⚖️ Balance Sheet", "📅 Monthly Analysis", "🤖 Scalpel AI"
    ])

    # TAB 1: Ledgers per account
    with tab1:
        st.subheader("View Detailed Ledger per Account")
        accounts = df[['account', 'account_description']].drop_duplicates().sort_values('account')
        account_options = {f"{row['account']} – {row['account_description']}": row['account'] for _, row in
                           accounts.iterrows()}
        selected_label = st.selectbox("Select account", list(account_options.keys()))
        selected_acc = account_options[selected_label]
        ledger_df = get_account_ledger(df, selected_acc)
        if not ledger_df.empty:
            st.dataframe(ledger_df, use_container_width=True)
            csv = ledger_df.to_csv(index=False)
            st.download_button("📥 Download this ledger (CSV)", csv, f"ledger_{selected_acc}.csv", "text/csv")

    # TAB 2: Trial Balance
    with tab2:
        st.subheader("Trial Balance")
        tb_display = tb[abs(tb['balance']) > 0.01].copy()
        st.dataframe(tb_display[['account', 'account_description', 'debit', 'credit', 'balance', 'category']],
                     use_container_width=True)

    # TAB 3: Detailed P&L (as required format)
    with tab3:
        st.subheader("Profit & Loss Statement")
        # Show revenue
        st.markdown("#### REVENUE")
        rev_display = rev_details.copy()
        rev_display.columns = ["Revenue Stream", "Amount (CAD)"]
        st.dataframe(rev_display, use_container_width=True)
        total_rev = rev_details['amount'].sum()
        st.markdown(f"**Total Revenue**  \n{total_rev:,.2f}  \n100.0%")

        st.markdown("#### COST OF GOODS SOLD")
        cogs_display = cogs_details.copy()
        cogs_display.columns = ["COGS Account", "Amount (CAD)"]
        st.dataframe(cogs_display, use_container_width=True)
        total_cogs = cogs_details['amount'].sum()
        st.markdown(f"**Total COGS**  \n{total_cogs:,.2f}  \n{total_cogs / total_rev * 100:.1f}%")

        gross_profit = total_rev - total_cogs
        st.markdown(f"**GROSS PROFIT**  \n{gross_profit:,.2f}  \n{gross_profit / total_rev * 100:.1f}%")

        st.markdown("#### OPERATING EXPENSES")
        # Group by subcategory
        for subcat in opex_details['subcategory'].unique():
            subcat_df = opex_details[opex_details['subcategory'] == subcat].copy()
            st.markdown(f"**▸ {subcat}**")
            sub_display = subcat_df[['account_description', 'amount']].copy()
            sub_display.columns = ["Expense", "Amount (CAD)"]
            st.dataframe(sub_display, use_container_width=True)
        total_opex = opex_details['amount'].sum()
        st.markdown(f"**Total Operating Expenses**  \n{total_opex:,.2f}  \n{total_opex / total_rev * 100:.1f}%")
        net_income_display = gross_profit - total_opex
        st.markdown(f"**NET INCOME**  \n{net_income_display:,.2f}  \n{net_income_display / total_rev * 100:.1f}%")

    # TAB 4: Balance Sheet (detailed)
    with tab4:
        st.subheader("Balance Sheet")
        st.markdown("#### ASSETS")
        st.dataframe(assets_details.rename(columns={"account_description": "Asset", "amount": "Amount"}),
                     use_container_width=True)
        st.markdown(f"**Total Assets**  \n{total_assets:,.2f}")
        st.markdown("#### LIABILITIES")
        st.dataframe(liabilities_details.rename(columns={"account_description": "Liability", "amount": "Amount"}),
                     use_container_width=True)
        st.markdown(f"**Total Liabilities**  \n{total_liabilities:,.2f}")
        st.markdown("#### EQUITY")
        st.dataframe(equity_details.rename(columns={"Account Description": "Equity", "Amount": "Amount"}),
                     use_container_width=True)
        st.markdown(f"**Total Equity**  \n{total_equity:,.2f}")
        st.markdown(f"**Liabilities + Equity**  \n{total_liabilities + total_equity:,.2f}")
        if abs(total_assets - (total_liabilities + total_equity)) < 0.01:
            st.success("✅ Balance Sheet balances.")
        else:
            st.warning(
                f"⚠️ Imbalance of {total_assets - (total_liabilities + total_equity):,.2f}. Check missing equity accounts.")

    # TAB 5: Monthly Analysis
    with tab5:
        if monthly is not None:
            st.subheader("Monthly Performance")
            st.dataframe(monthly, use_container_width=True)
            fig = px.line(monthly, x='year_month', y=['Revenue', 'Net Income'], title="Revenue & Net Income Over Time")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No effective date column found. Monthly analysis requires a date column.")

    # TAB 6: AI Chat (placeholder – can be added later)
    with tab6:
        st.subheader("Ask Scalpel AI")
        st.info("AI chat requires API keys (Groq or Anthropic). The core accounting engine works without AI.")
        # We keep a simplified version that only uses the data.
        user_question = st.text_input("Your question (prototype):", "What is our gross margin?")
        if st.button("Analyze (Mock)"):
            context = f"Revenue: {total_rev:,.2f}, COGS: {total_cogs:,.2f}, OpEx: {total_opex:,.2f}, Net Income: {net_income_display:,.2f}"
            st.write(
                f"**Scalpel AI (mock):** Based on the data, gross margin is {gross_profit / total_rev * 100:.1f}%.")

    # -------------------------------------------------------------------
    # DOWNLOAD BUTTON
    # -------------------------------------------------------------------
    st.markdown("---")
    excel_data = download_detailed_excel(tb, pl_summary, rev_details, cogs_details, opex_details, assets_details,
                                         liabilities_details, equity_details, monthly)
    st.download_button(
        label="📥 Download Full Financials (Excel)",
        data=excel_data,
        file_name="scalpel_detailed_financials.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    main()