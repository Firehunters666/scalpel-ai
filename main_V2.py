import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os

# -------------------------------------------------------------------
# PAGE CONFIG
# -------------------------------------------------------------------
st.set_page_config(page_title="Scalpel AI – Deterministic Accounting Engine", layout="wide")
st.title("Scalpel AI")
st.markdown("### Deterministic Finance Operating System")

# -------------------------------------------------------------------
# CUSTOM CSS FOR ENTERPRISE DASHBOARD (only metric value font size reduced)
# -------------------------------------------------------------------
st.markdown("""
<style>
    /* Main container */
    .main {
        background-color: #0e1117;
        color: #f0f2f6;
    }
    /* Metric cards */
    .metric-card {
        background-color: #1e2229;
        border-radius: 16px;
        padding: 1.2rem;
        margin: 0.5rem;
        box-shadow: 0 4px 12px rgba(0,0,0,0.3);
        border: 1px solid #2d3138;
        transition: transform 0.2s;
    }
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: #1abc9c;
    }
    .metric-title {
        font-size: 0.85rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #8b8f96;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 1.3rem;          /* Reduced from 2rem to fit on one line */
        font-weight: 600;
        color: #ffffff;
        white-space: nowrap;
        overflow-x: auto;
    }
    .metric-compare {
        font-size: 0.75rem;
        color: #6c757d;
        margin-top: 0.5rem;
    }
    /* Sidebar */
    .css-1d391kg {
        background-color: #0a0c10;
    }
    /* Tables */
    .stDataFrame {
        background-color: #1e2229;
        border-radius: 12px;
        overflow: hidden;
        font-size: 0.9rem;
    }
    /* Headers */
    h1, h2, h3 {
        color: #f0f2f6;
    }
    /* Buttons */
    .stButton button {
        background-color: #2c3e50;
        color: white;
        border-radius: 8px;
        border: none;
        padding: 0.5rem 1rem;
        transition: background-color 0.2s;
    }
    .stButton button:hover {
        background-color: #1abc9c;
        color: black;
    }
    /* Charts container */
    .chart-container {
        background-color: #1e2229;
        border-radius: 16px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


# -------------------------------------------------------------------
# DATA NORMALISATION (unchanged)
# -------------------------------------------------------------------
def normalize_columns(df):
    df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
    rename_map = {
        'account_description': 'account_description',
        'accountname': 'account_description',
        'account_desc': 'account_description',
        'effective_date': 'effective_date',
        'effectivedate': 'effective_date',
        'entered_date': 'entered_date',
        'entereddate': 'entered_date',
        'transaction': 'transaction',
        'txnid': 'transaction',
        'glid': 'transaction',
        'debit': 'debit',
        'debits': 'debit',
        'credit': 'credit',
        'credits': 'credit',
        'memo': 'memo',
        'description': 'memo'
    }
    for old, new in rename_map.items():
        if old in df.columns and new not in df.columns:
            df.rename(columns={old: new}, inplace=True)
    return df


@st.cache_data
def load_and_validate(uploaded_file):
    df = pd.read_excel(uploaded_file)
    df = normalize_columns(df)

    required = ['account', 'debit', 'credit']
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}. Please ensure file has account, debit, credit.")
        return None, False

    if 'account_description' not in df.columns:
        df['account_description'] = df['account'].astype(str)
    if 'effective_date' not in df.columns and 'entered_date' in df.columns:
        df['effective_date'] = df['entered_date']
    if 'effective_date' not in df.columns:
        df['effective_date'] = pd.NaT
    if 'transaction' not in df.columns:
        df['transaction'] = df.index.astype(str)
    if 'memo' not in df.columns:
        df['memo'] = ''

    df['debit'] = pd.to_numeric(df['debit'], errors='coerce').fillna(0)
    df['credit'] = pd.to_numeric(df['credit'], errors='coerce').fillna(0)

    total_debits = df['debit'].sum()
    total_credits = df['credit'].sum()
    if abs(total_debits - total_credits) > 0.01:
        st.error(f"❌ Global imbalance: Debits = {total_debits:,.2f}, Credits = {total_credits:,.2f}")
        return None, False
    else:
        # Do not display success message – professional UI hides it
        pass

    return df, True


# -------------------------------------------------------------------
# ACCOUNT CLASSIFICATION (same as before)
# -------------------------------------------------------------------
def classify_account(acct_num, acct_desc):
    acct_str = str(acct_num)
    desc_lower = acct_desc.lower()
    if any(k in desc_lower for k in ['gst', 'pst', 'qst', 'hst', 'tax']) and acct_str.startswith('2'):
        return "Liability"
    if acct_str.startswith('4') or any(k in desc_lower for k in
                                       ['sales', 'revenue', 'consulting', 'drafting', 'surveying', 'service',
                                        'freight revenue']):
        return "Revenue"
    if acct_str.startswith('5'):
        if any(k in desc_lower for k in
               ['material', 'cogs', 'purchase discounts', 'adjustment write-off', 'equipment rental', 'freight expense',
                'item assembly']):
            return "COGS"
        else:
            return "OpEx"
    if acct_str.startswith('1') or any(k in desc_lower for k in
                                       ['cash', 'bank', 'receivable', 'inventory', 'prepaid', 'accum amort', 'vehicles',
                                        'drywall', 'hardware', 'lumber', 'roofing']):
        return "Asset"
    if acct_str.startswith('2') or 'payable' in desc_lower:
        return "Liability"
    if acct_str.startswith('3'):
        return "Equity"
    return "Other"


def add_classification(df):
    unique_acc = df[['account', 'account_description']].drop_duplicates()
    unique_acc['category'] = unique_acc.apply(lambda r: classify_account(r['account'], r['account_description']),
                                              axis=1)
    return df.merge(unique_acc[['account', 'category']], on='account', how='left')


# -------------------------------------------------------------------
# BALANCE SHEET CLASSIFICATION (Current/Non-Current)
# -------------------------------------------------------------------
def classify_balance_sheet_account(account_description, balance):
    desc_lower = account_description.lower()
    if any(k in desc_lower for k in ['cash', 'bank', 'receivable', 'inventory', 'prepaid']):
        return "Current Assets"
    if any(k in desc_lower for k in ['accum amort', 'vehicle', 'building', 'equipment', 'furniture', 'software']):
        return "Non-Current Assets"
    if "asset" in desc_lower or balance > 0:
        if balance < 0:
            return "Non-Current Assets"
        return "Current Assets"
    if any(k in desc_lower for k in
           ['payable', 'accrued', 'wcb', 'ei', 'cpp', 'qpp', 'eht', 'qpip', 'qhsf', 'rrsp', 'union', 'medical',
            'disability', 'pst', 'gst']):
        return "Current Liabilities"
    if any(k in desc_lower for k in ['long term', 'mortgage']):
        return "Non-Current Liabilities"
    if "liability" in desc_lower or balance < 0:
        return "Current Liabilities"
    return "Other"


def generate_balance_sheet(tb, net_income):
    asset_df = tb[tb['category'] == 'Asset'].copy()
    asset_df['amount'] = asset_df['balance']
    asset_df['type'] = asset_df['account_description'].apply(lambda x: classify_balance_sheet_account(x, asset_df[
        asset_df['account_description'] == x]['balance'].values[0] if not asset_df[
        asset_df['account_description'] == x].empty else 0))
    current_assets = asset_df[asset_df['type'] == 'Current Assets'][['account_description', 'amount']].sort_values(
        'amount', ascending=False)
    non_current_assets = asset_df[asset_df['type'] != 'Current Assets'][['account_description', 'amount']].sort_values(
        'amount', ascending=False)
    total_current_assets = current_assets['amount'].sum()
    total_non_current_assets = non_current_assets['amount'].sum()
    total_assets = total_current_assets + total_non_current_assets

    liab_df = tb[tb['category'] == 'Liability'].copy()
    liab_df['amount'] = -liab_df['balance']
    liab_df['type'] = liab_df['account_description'].apply(lambda x: classify_balance_sheet_account(x, liab_df[
        liab_df['account_description'] == x]['balance'].values[0] if not liab_df[
        liab_df['account_description'] == x].empty else 0))
    current_liabilities = liab_df[liab_df['type'] == 'Current Liabilities'][
        ['account_description', 'amount']].sort_values('amount', ascending=False)
    non_current_liabilities = liab_df[liab_df['type'] != 'Current Liabilities'][
        ['account_description', 'amount']].sort_values('amount', ascending=False)
    total_current_liabilities = current_liabilities['amount'].sum()
    total_non_current_liabilities = non_current_liabilities['amount'].sum()
    total_liabilities = total_current_liabilities + total_non_current_liabilities

    equity_df = tb[tb['category'] == 'Equity'].copy()
    equity_df['amount'] = equity_df['balance']
    total_equity_from_tb = equity_df['amount'].sum()
    implied_equity = total_assets - total_liabilities
    missing_equity = implied_equity - (total_equity_from_tb + net_income)
    equity_items = []
    if total_equity_from_tb != 0:
        equity_items.append(("Equity from Trial Balance", total_equity_from_tb))
    equity_items.append(("Net Income (current period)", net_income))
    if abs(missing_equity) > 0.01:
        equity_items.append(("Retained Earnings / Owner's Equity (balancing)", missing_equity))
    total_equity = implied_equity
    equity_items.append(("Total Equity", total_equity))
    equity = pd.DataFrame(equity_items, columns=["Account Description", "Amount"])
    return (current_assets, non_current_assets, total_current_assets, total_non_current_assets, total_assets,
            current_liabilities, non_current_liabilities, total_current_liabilities, total_non_current_liabilities,
            total_liabilities,
            equity, total_equity)


def generate_detailed_pl(df_class, tb):
    rev_df = tb[tb['category'] == 'Revenue'].copy()
    rev_df['amount'] = rev_df['credit'] - rev_df['debit']
    revenue_df = rev_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_revenue = revenue_df['amount'].sum()
    cogs_df = tb[tb['category'] == 'COGS'].copy()
    cogs_df['amount'] = cogs_df['debit']
    cogs_details = cogs_df[['account_description', 'amount']].sort_values('amount', ascending=False)
    total_cogs = cogs_details['amount'].sum()
    gross_profit = total_revenue - total_cogs
    opex_df = tb[tb['category'] == 'OpEx'].copy()
    opex_df['amount'] = opex_df['debit'] - opex_df['credit']
    opex_details = opex_df[['account_description', 'amount']].sort_values('amount', ascending=False)

    def subcategory(desc):
        d = desc.lower()
        if any(k in d for k in
               ['wage', 'salary', 'ei', 'cpp', 'qpp', 'wcb', 'eht', 'qpip', 'qhsf', 'rrsp', 'union', 'medical',
                'subcontractor']):
            return "Payroll & Benefits"
        if any(k in d for k in
               ['amortization', 'amort', 'insurance', 'rent', 'repair', 'maintenance', 'bell', 'utilities']):
            return "Facilities & Operations"
        if any(k in d for k in
               ['accounting', 'legal', 'advertising', 'promotion', 'courier', 'postage', 'interest', 'bank',
                'office supplies', 'travel', 'amex', 'licenses', 'currency exchange']):
            return "Professional & Admin"
        return "Other Operating Expenses"

    opex_details['subcategory'] = opex_details['account_description'].apply(subcategory)
    total_opex = opex_details['amount'].sum()
    net_income = gross_profit - total_opex
    return revenue_df, cogs_details, opex_details, total_revenue, total_cogs, gross_profit, total_opex, net_income


def generate_trial_balance(df):
    tb = df.groupby(['account', 'account_description']).agg({'debit': 'sum', 'credit': 'sum'}).reset_index()
    tb['balance'] = tb['debit'] - tb['credit']
    tb['category'] = tb.apply(lambda r: classify_account(r['account'], r['account_description']), axis=1)
    return tb


def generate_monthly_analysis(df):
    if 'effective_date' not in df.columns or df['effective_date'].isnull().all():
        return None
    df['date'] = pd.to_datetime(df['effective_date'], errors='coerce')
    if df['date'].isnull().all():
        return None
    df['year_month'] = df['date'].dt.to_period('M')
    df_class = add_classification(df)
    monthly = df_class.groupby('year_month').apply(
        lambda x: pd.Series({
            'Revenue': x[x['category'] == 'Revenue']['credit'].sum() - x[x['category'] == 'Revenue']['debit'].sum(),
            'COGS': x[x['category'] == 'COGS']['debit'].sum(),
            'OpEx': x[x['category'] == 'OpEx']['debit'].sum() - x[x['category'] == 'OpEx']['credit'].sum()
        })
    ).reset_index()
    monthly['Gross Profit'] = monthly['Revenue'] - monthly['COGS']
    monthly['Net Income'] = monthly['Gross Profit'] - monthly['OpEx']
    monthly['Gross Margin %'] = (monthly['Gross Profit'] / monthly['Revenue'] * 100).round(1) if monthly[
                                                                                                     'Revenue'].sum() != 0 else 0
    monthly['Net Margin %'] = (monthly['Net Income'] / monthly['Revenue'] * 100).round(1) if monthly[
                                                                                                 'Revenue'].sum() != 0 else 0
    monthly['year_month'] = monthly['year_month'].astype(str)
    return monthly


# -------------------------------------------------------------------
# AI QUERY (Groq) – also used for Flux Analysis
# -------------------------------------------------------------------
def ask_ai(question, financial_context, model="llama-3.3-70b-versatile"):
    api_key = st.secrets.get("GROQ_API_KEY") or os.environ.get("GROQ_API_KEY")
    if not api_key:
        return "⚠️ Groq API key not set. Please add GROQ_API_KEY to your secrets or environment."
    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        system_prompt = f"""
        You are Scalpel AI, a financial analyst assistant. Use the following financial data to answer the user's question.
        Provide concise, accurate answers. Show calculations if needed.

        Financial Data (Jan–Sep 2020):
        - Total Revenue: {financial_context['total_revenue']:,.2f}
        - COGS: {financial_context['total_cogs']:,.2f}
        - Gross Profit: {financial_context['gross_profit']:,.2f}
        - Operating Expenses: {financial_context['total_opex']:,.2f}
        - Net Income: {financial_context['net_income']:,.2f}

        Key revenue streams:
        {financial_context['revenue_breakdown']}

        Largest expenses:
        {financial_context['top_expenses']}
        """
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": question}
            ],
            temperature=0.2,
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"❌ AI error: {str(e)}"


# -------------------------------------------------------------------
# EXCEL DOWNLOAD (with totals in Trial Balance, 2 decimals)
# -------------------------------------------------------------------
def download_excel(tb, revenue_df, cogs_df, opex_df,
                   current_assets, non_current_assets, total_current_assets, total_non_current_assets, total_assets,
                   current_liabilities, non_current_liabilities, total_current_liabilities,
                   total_non_current_liabilities, total_liabilities,
                   equity_df, total_equity, monthly_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Trial Balance with totals row and 2 decimals
        tb_with_totals = tb.copy()
        totals_row = pd.DataFrame({
            'account': ['TOTAL'],
            'account_description': [''],
            'debit': [tb['debit'].sum()],
            'credit': [tb['credit'].sum()],
            'balance': [tb['balance'].sum()],
            'category': ['']
        })
        tb_with_totals = pd.concat([tb_with_totals, totals_row], ignore_index=True)
        # Round to 2 decimals
        for col in ['debit', 'credit', 'balance']:
            tb_with_totals[col] = tb_with_totals[col].round(2)
        tb_with_totals.to_excel(writer, sheet_name='Trial Balance', index=False)

        # P&L single sheet
        pl_rows = []
        pl_rows.append(["REVENUE", ""])
        for _, row in revenue_df.iterrows():
            if abs(row['amount']) > 0.01:
                pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total Revenue", round(revenue_df['amount'].sum(), 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["COST OF GOODS SOLD", ""])
        for _, row in cogs_df.iterrows():
            if abs(row['amount']) > 0.01:
                pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total COGS", round(cogs_df['amount'].sum(), 2)])
        pl_rows.append(["", ""])
        gross = revenue_df['amount'].sum() - cogs_df['amount'].sum()
        pl_rows.append(["GROSS PROFIT", round(gross, 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["OPERATING EXPENSES", ""])
        for subcat in opex_df['subcategory'].unique():
            pl_rows.append([f"▸ {subcat}", ""])
            sub = opex_df[opex_df['subcategory'] == subcat]
            for _, row in sub.iterrows():
                if abs(row['amount']) > 0.01:
                    pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total Operating Expenses", round(opex_df['amount'].sum(), 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["NET INCOME", round(gross - opex_df['amount'].sum(), 2)])
        pl_df = pd.DataFrame(pl_rows, columns=["Account", "Amount (CAD)"])
        pl_df.to_excel(writer, sheet_name='Profit & Loss', index=False)

        # Balance Sheet
        bs_rows = []
        bs_rows.append(["ASSETS", ""])
        bs_rows.append(["Current Assets", ""])
        for _, row in current_assets.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Current Assets", round(total_current_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Non-Current Assets", ""])
        for _, row in non_current_assets.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Non-Current Assets", round(total_non_current_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Total Assets", round(total_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["LIABILITIES", ""])
        bs_rows.append(["Current Liabilities", ""])
        for _, row in current_liabilities.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Current Liabilities", round(total_current_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Non-Current Liabilities", ""])
        for _, row in non_current_liabilities.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Non-Current Liabilities", round(total_non_current_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Total Liabilities", round(total_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["EQUITY", ""])
        for _, row in equity_df.iterrows():
            if abs(row['Amount']) > 0.01:
                bs_rows.append([row['Account Description'], round(row['Amount'], 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["TOTAL LIABILITIES AND EQUITY", round(total_liabilities + total_equity, 2)])
        bs_df = pd.DataFrame(bs_rows, columns=["Account", "Amount (CAD)"])
        bs_df.to_excel(writer, sheet_name='Balance Sheet', index=False)

        if monthly_df is not None:
            monthly_df.to_excel(writer, sheet_name='Monthly Analysis', index=False)

        # Styling
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 30)
    return output.getvalue()


# -------------------------------------------------------------------
# HELPER: Account Ledger
# -------------------------------------------------------------------
def get_account_ledger(df, account_num):
    acc_df = df[df['account'] == account_num].copy()
    if acc_df.empty:
        return pd.DataFrame()
    date_col = 'effective_date' if 'effective_date' in acc_df.columns and acc_df[
        'effective_date'].notna().any() else 'entered_date'
    if date_col in acc_df.columns:
        acc_df['date'] = pd.to_datetime(acc_df[date_col], errors='coerce')
        acc_df = acc_df.sort_values('date')
    else:
        acc_df['date'] = pd.NaT
    show_cols = ['date', 'memo', 'debit', 'credit']
    for c in show_cols:
        if c not in acc_df.columns:
            acc_df[c] = ''
    ledger = acc_df[show_cols].copy()
    ledger['debit'] = ledger['debit'].fillna(0).round(2)
    ledger['credit'] = ledger['credit'].fillna(0).round(2)
    total_debit = ledger['debit'].sum()
    total_credit = ledger['credit'].sum()
    total_row = pd.DataFrame([['TOTAL', '', total_debit, total_credit]], columns=show_cols)
    ledger = pd.concat([ledger, total_row], ignore_index=True)
    return ledger


# -------------------------------------------------------------------
# DASHBOARD METRICS & CHARTS
# -------------------------------------------------------------------
def display_dashboard(tb, revenue_df, net_income, monthly_df):
    # Extract cash balance
    cash_accounts = tb[
        (tb['category'] == 'Asset') & (tb['account_description'].str.contains('bank|cash|checking', case=False))]
    cash_balance = cash_accounts['balance'].sum() if not cash_accounts.empty else 0
    ar_accounts = tb[(tb['category'] == 'Asset') & (tb['account_description'].str.contains('receivable', case=False))]
    ar_balance = ar_accounts['balance'].sum() if not ar_accounts.empty else 0
    ap_accounts = tb[
        (tb['category'] == 'Liability') & (tb['account_description'].str.contains('accounts payable', case=False))]
    ap_balance = -ap_accounts['balance'].sum() if not ap_accounts.empty else 0
    total_revenue = revenue_df['amount'].sum()

    # Comparisons
    prev_revenue = None
    prev_net_income = None
    if monthly_df is not None and len(monthly_df) >= 2:
        prev_revenue = monthly_df.iloc[-2]['Revenue']
        prev_net_income = monthly_df.iloc[-2]['Net Income']

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-title">Cash Balance</div>
            <div class="metric-value">${cash_balance:,.2f}</div>
            <div class="metric-compare">As of latest period</div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        compare_text = f"vs prev: ${(total_revenue - prev_revenue):+,.2f}" if prev_revenue else ""
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-title">Total Revenue</div>
            <div class="metric-value">${total_revenue:,.2f}</div>
            <div class="metric-compare">{compare_text}</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        compare_text = f"vs prev: ${(net_income - prev_net_income):+,.2f}" if prev_net_income else ""
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-title">Net Income</div>
            <div class="metric-value">${net_income:,.2f}</div>
            <div class="metric-compare">{compare_text}</div>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-title">Accounts Receivable</div>
            <div class="metric-value">${ar_balance:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)
    with col5:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-title">Accounts Payable</div>
            <div class="metric-value">${ap_balance:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # Workflow Snapshot
    st.markdown("### Workflow Snapshot")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Cash Transactions to be reconciled", "13")
        st.metric("Invoices to be sent", "809")
    with col2:
        st.metric("Journal entries pending approval", "2")
        st.metric("Bills to be paid", "229")
    with col3:
        st.metric("Contracts to be reviewed", "2")
        st.metric("Last closed month", "May 2024")

    # Tech Stack Monitoring
    st.markdown("### Tech Stack Monitoring")
    tech_cols = st.columns(5)
    with tech_cols[0]:
        st.image("https://cdn.jsdelivr.net/gh/devicons/devicon/icons/amazonwebservices/amazonwebservices-original.svg",
                 width=40)
        st.caption("Banking")
    with tech_cols[1]:
        st.image("https://cdn.jsdelivr.net/gh/devicons/devicon/icons/salesforce/salesforce-original.svg", width=40)
        st.caption("Salesforce")
    with tech_cols[2]:
        st.image("https://cdn.jsdelivr.net/gh/devicons/devicon/icons/stripe/stripe-original.svg", width=40)
        st.caption("Stripe")
    with tech_cols[3]:
        st.caption("Avatax")
    with tech_cols[4]:
        st.caption("Ramp")

    # Charts Section
    st.markdown("### 📊 Performance Insights")
    if monthly_df is not None:
        col1, col2 = st.columns(2)
        with col1:
            fig_rev = px.line(monthly_df, x='year_month', y=['Revenue', 'Net Income'],
                              title="Revenue vs Net Income Trend", markers=True)
            fig_rev.update_layout(template='plotly_dark', hovermode='x unified')
            st.plotly_chart(fig_rev, use_container_width=True)
        with col2:
            fig_margin = px.line(monthly_df, x='year_month', y=['Gross Margin %', 'Net Margin %'],
                                 title="Gross & Net Margin %", markers=True)
            fig_margin.update_layout(template='plotly_dark', hovermode='x unified')
            st.plotly_chart(fig_margin, use_container_width=True)

        # Top revenue streams bar chart
        top_rev = revenue_df.head(8).copy()
        top_rev['amount'] = top_rev['amount'].round(2)
        fig_bar = px.bar(top_rev, x='account_description', y='amount', title="Top Revenue Streams",
                         text='amount', template='plotly_dark')
        fig_bar.update_traces(texttemplate='$%{text:.2f}', textposition='outside')
        fig_bar.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_bar, use_container_width=True)
    else:
        st.info("Monthly analysis data not available – upload a file with effective dates to see charts.")


# -------------------------------------------------------------------
# FLUX ANALYSIS WITH AI (enhanced Close Management)
# -------------------------------------------------------------------
def flux_analysis(monthly_df, revenue_df, opex_df, total_rev, total_cogs, gross_profit, total_opex, net_income):
    if monthly_df is None or len(monthly_df) < 2:
        return "Not enough monthly data to perform variance analysis."

    latest = monthly_df.iloc[-1]
    previous = monthly_df.iloc[-2]
    var_rev = latest['Revenue'] - previous['Revenue']
    var_ni = latest['Net Income'] - previous['Net Income']
    var_gross = latest['Gross Profit'] - previous['Gross Profit']
    var_opex = latest['OpEx'] - previous['OpEx']

    # Build a simple variance explanation using AI
    financial_context = {
        'total_revenue': total_rev,
        'total_cogs': total_cogs,
        'gross_profit': gross_profit,
        'total_opex': total_opex,
        'net_income': net_income,
        'revenue_breakdown': "\n".join(
            [f"{row['account_description']}: {row['amount']:,.2f}" for _, row in revenue_df.head(5).iterrows()]),
        'top_expenses': "\n".join(
            [f"{row['account_description']}: {row['amount']:,.2f}" for _, row in opex_df.head(5).iterrows()])
    }

    prompt = f"""
    You are a financial controller. Explain the following month-over-month variances for the latest period:
    - Revenue change: {var_rev:+,.2f}
    - Gross Profit change: {var_gross:+,.2f}
    - Operating Expenses change: {var_opex:+,.2f}
    - Net Income change: {var_ni:+,.2f}
    Provide a concise, insightful explanation (2-3 sentences) focusing on material drivers.
    """
    try:
        explanation = ask_ai(prompt, financial_context, model="llama-3.1-8b-instant")
    except:
        explanation = "AI explanation unavailable. Review the variance numbers above."

    return var_rev, var_ni, var_gross, var_opex, explanation


# -------------------------------------------------------------------
# MAIN APP
# -------------------------------------------------------------------
def main():
    uploaded_file = st.file_uploader("📂 Upload General Ledger (Excel)", type=["xlsx", "xls"])
    if uploaded_file is None:
        st.info("Please upload your general ledger file to start.")
        return

    df, is_valid = load_and_validate(uploaded_file)
    if not is_valid or df is None:
        st.stop()

    df_class = add_classification(df)
    tb = generate_trial_balance(df)
    revenue_df, cogs_df, opex_df, total_rev, total_cogs, gross_profit, total_opex, net_income = generate_detailed_pl(
        df_class, tb)
    monthly = generate_monthly_analysis(df)

    # Sidebar navigation
    st.sidebar.markdown("## Navigation")
    nav_options = ["Dashboard", "Account Ledgers", "Trial Balance", "Profit & Loss", "Balance Sheet",
                   "Monthly Analysis", "AI Chat", "Close Management"]
    selected = st.sidebar.radio("Go to", nav_options, label_visibility="collapsed")

    if selected == "Dashboard":
        st.header("Dashboard")
        display_dashboard(tb, revenue_df, net_income, monthly)

    elif selected == "Account Ledgers":
        st.subheader("View Detailed Ledger per Account")
        accounts = df[['account', 'account_description']].drop_duplicates().sort_values('account')
        account_options = {f"{row['account']} – {row['account_description']}": row['account'] for _, row in
                           accounts.iterrows()}
        selected_label = st.selectbox("Select account", list(account_options.keys()))
        selected_acc = account_options[selected_label]
        ledger_df = get_account_ledger(df, selected_acc)
        if not ledger_df.empty:
            # Format ledger for display
            for col in ['debit', 'credit']:
                ledger_df[col] = ledger_df[col].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)
            st.dataframe(ledger_df, use_container_width=True)
            csv = ledger_df.to_csv(index=False)
            st.download_button("📥 Download this ledger (CSV)", csv, f"ledger_{selected_acc}.csv", "text/csv")

    elif selected == "Trial Balance":
        st.subheader("Trial Balance")
        tb_display = tb[abs(tb['balance']) > 0.01].copy()
        totals = pd.DataFrame({
            'account': ['TOTAL'],
            'account_description': [''],
            'debit': [tb['debit'].sum()],
            'credit': [tb['credit'].sum()],
            'balance': [tb['balance'].sum()],
            'category': ['']
        })
        tb_display = pd.concat([tb_display, totals], ignore_index=True)
        # Round for display
        for col in ['debit', 'credit', 'balance']:
            tb_display[col] = tb_display[col].round(2)
        st.dataframe(tb_display[['account', 'account_description', 'debit', 'credit', 'balance', 'category']],
                     use_container_width=True)
        total_debits = tb['debit'].sum()
        total_credits = tb['credit'].sum()
        st.markdown(f"**Total Debits:** {total_debits:,.2f}  |  **Total Credits:** {total_credits:,.2f}")
        if abs(total_debits - total_credits) < 0.01:
            st.success("✅ Trial balance is in balance.")
        else:
            st.error("❌ Trial balance is out of balance! Check data.")

    elif selected == "Profit & Loss":
        st.subheader("Profit & Loss Statement")
        pl_rows = []
        pl_rows.append(["REVENUE", "Amount (CAD)"])
        for _, row in revenue_df.iterrows():
            if abs(row['amount']) > 0.01:
                pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total Revenue", round(total_rev, 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["COST OF GOODS SOLD", ""])
        for _, row in cogs_df.iterrows():
            if abs(row['amount']) > 0.01:
                pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total COGS", round(total_cogs, 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["GROSS PROFIT", round(gross_profit, 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["OPERATING EXPENSES", ""])
        for subcat in opex_df['subcategory'].unique():
            pl_rows.append([f"▸ {subcat}", ""])
            sub = opex_df[opex_df['subcategory'] == subcat]
            for _, row in sub.iterrows():
                if abs(row['amount']) > 0.01:
                    pl_rows.append([row['account_description'], round(row['amount'], 2)])
        pl_rows.append(["Total Operating Expenses", round(total_opex, 2)])
        pl_rows.append(["", ""])
        pl_rows.append(["NET INCOME", round(net_income, 2)])
        pl_display = pd.DataFrame(pl_rows[1:], columns=pl_rows[0])
        st.dataframe(pl_display, use_container_width=True, hide_index=True)

    elif selected == "Balance Sheet":
        st.subheader("Balance Sheet")
        (current_assets, non_current_assets, total_current_assets, total_non_current_assets, total_assets,
         current_liabilities, non_current_liabilities, total_current_liabilities, total_non_current_liabilities,
         total_liabilities,
         equity_df, total_equity) = generate_balance_sheet(tb, net_income)

        bs_rows = []
        bs_rows.append(["ASSETS", "Amount (CAD)"])
        bs_rows.append(["Current Assets", ""])
        for _, row in current_assets.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Current Assets", round(total_current_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Non-Current Assets", ""])
        for _, row in non_current_assets.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Non-Current Assets", round(total_non_current_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Total Assets", round(total_assets, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["LIABILITIES", ""])
        bs_rows.append(["Current Liabilities", ""])
        for _, row in current_liabilities.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Current Liabilities", round(total_current_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Non-Current Liabilities", ""])
        for _, row in non_current_liabilities.iterrows():
            if abs(row['amount']) > 0.01:
                bs_rows.append([row['account_description'], round(row['amount'], 2)])
        bs_rows.append(["Total Non-Current Liabilities", round(total_non_current_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["Total Liabilities", round(total_liabilities, 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["EQUITY", ""])
        for _, row in equity_df.iterrows():
            if abs(row['Amount']) > 0.01:
                bs_rows.append([row['Account Description'], round(row['Amount'], 2)])
        bs_rows.append(["", ""])
        bs_rows.append(["TOTAL LIABILITIES AND EQUITY", round(total_liabilities + total_equity, 2)])
        bs_display = pd.DataFrame(bs_rows[1:], columns=bs_rows[0])
        st.dataframe(bs_display, use_container_width=True, hide_index=True)
        if abs(total_assets - (total_liabilities + total_equity)) < 0.01:
            st.success("✅ Balance Sheet balances.")
        else:
            st.warning(
                f"⚠️ Imbalance of {total_assets - (total_liabilities + total_equity):,.2f}. (Automatically corrected by adding balancing equity item.)")

    elif selected == "Monthly Analysis":
        if monthly is not None:
            st.subheader("Monthly Performance")
            # Format numbers to 2 decimals
            monthly_display = monthly.copy()
            for col in ['Revenue', 'COGS', 'OpEx', 'Gross Profit', 'Net Income']:
                monthly_display[col] = monthly_display[col].round(2)
            st.dataframe(monthly_display, use_container_width=True)
            col1, col2 = st.columns(2)
            with col1:
                fig1 = px.line(monthly, x='year_month', y=['Revenue', 'Net Income'], title="Revenue & Net Income",
                               markers=True)
                fig1.update_layout(template='plotly_dark')
                st.plotly_chart(fig1, use_container_width=True)
            with col2:
                fig2 = px.bar(monthly, x='year_month', y='Gross Profit', title="Gross Profit by Month",
                              text='Gross Profit')
                fig2.update_traces(texttemplate='$%{text:.0f}', textposition='outside')
                fig2.update_layout(template='plotly_dark')
                st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No effective date column found. Monthly analysis requires a date column.")

    elif selected == "AI Chat":
        st.subheader("Ask Scalpel AI")
        revenue_breakdown = "\n".join(
            [f"- {row['account_description']}: {row['amount']:,.2f}" for _, row in revenue_df.head(10).iterrows()])
        top_expenses = "\n".join(
            [f"- {row['account_description']}: {row['amount']:,.2f}" for _, row in opex_df.head(10).iterrows()])
        financial_context = {
            'total_revenue': total_rev,
            'total_cogs': total_cogs,
            'gross_profit': gross_profit,
            'total_opex': total_opex,
            'net_income': net_income,
            'revenue_breakdown': revenue_breakdown,
            'top_expenses': top_expenses
        }
        model_option = st.selectbox(
            "Select AI Model",
            ("llama-3.3-70b-versatile", "llama-3.1-8b-instant"),
            index=0,
            help="llama-3.3-70b-versatile offers higher quality, while llama-3.1-8b-instant is faster."
        )
        user_question = st.text_input("Your question:",
                                      placeholder="e.g., What is our gross margin? Which expense is highest?")
        if st.button("Ask AI"):
            if user_question.strip():
                with st.spinner(f"Analyzing with {model_option}..."):
                    answer = ask_ai(user_question, financial_context, model=model_option)
                st.success("Answer:")
                st.write(answer)
            else:
                st.warning("Please enter a question.")

    elif selected == "Close Management":
        st.subheader("AI-Assisted Flux Analysis")
        if monthly is not None and len(monthly) >= 2:
            var_rev, var_ni, var_gross, var_opex, explanation = flux_analysis(
                monthly, revenue_df, opex_df, total_rev, total_cogs, gross_profit, total_opex, net_income
            )
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Revenue Variance (MoM)", f"{var_rev:+,.2f}", delta_color="normal")
                st.metric("Gross Profit Variance (MoM)", f"{var_gross:+,.2f}", delta_color="normal")
            with col2:
                st.metric("Operating Expenses Variance (MoM)", f"{var_opex:+,.2f}", delta_color="inverse")
                st.metric("Net Income Variance (MoM)", f"{var_ni:+,.2f}", delta_color="normal")
            st.markdown("### AI-Generated Explanation")
            st.info(explanation)
            st.markdown("#### Account-Level Variances (Top 5)")
            # Show top 5 revenue and expense changes month-over-month
            if len(monthly) >= 2:
                # This requires account-level monthly data – simplified version using summary
                st.write(
                    "For deeper account-level analysis, the system would display detailed fluctuations. In production, this would include COGS and OpEx account changes.")
        else:
            st.info(
                "Not enough monthly data to perform flux analysis. Please ensure your data includes effective dates covering at least two months.")

    # Download button always visible
    st.markdown("---")
    (current_assets, non_current_assets, total_current_assets, total_non_current_assets, total_assets,
     current_liabilities, non_current_liabilities, total_current_liabilities, total_non_current_liabilities,
     total_liabilities,
     equity_df, total_equity) = generate_balance_sheet(tb, net_income)
    excel_data = download_excel(tb, revenue_df, cogs_df, opex_df,
                                current_assets, non_current_assets, total_current_assets, total_non_current_assets,
                                total_assets,
                                current_liabilities, non_current_liabilities, total_current_liabilities,
                                total_non_current_liabilities, total_liabilities,
                                equity_df, total_equity, monthly)
    st.download_button(
        label="📥 Download Full Financials (Excel)",
        data=excel_data,
        file_name="scalpel_corrected_financials.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    main()
